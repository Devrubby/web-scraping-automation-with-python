from pathlib import  Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import  get_column_letter
output_path = Path("Output_Dir")


def style_columns(sheet):
    # Bold the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    # Auto-size each column
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 5



number_format = '$#,##0.00'

def build_excel_report(books_df, stats,
                       ratings, output_path):
    wkb = Workbook()
    ws1 = wkb.active
    ws1.title = "Books"
    ws1.append(list(books_df.columns))
    for row in list(books_df.values.tolist()):
        ws1.append(row)
    for cell in ws1["B"][1:]:
        cell.number_format = number_format
    style_columns(ws1)

    ws2 = wkb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    for key, value in stats.items():
        ws2.append([key, value])
    for cell in ws2["B"][2:5]:
        cell.number_format = number_format
    style_columns(ws2)

    ws3 = wkb.create_sheet("Ratings Distribution")
    ws3.append(["Rating", "Count"])
    for key, value in ratings.items():
        ws3.append([key, value])
    style_columns(ws3)
    wkb.save(output_path / "reports.xlsx")